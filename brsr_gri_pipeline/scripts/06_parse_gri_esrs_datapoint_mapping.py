"""
Parses the official GRI-ESRS Data Point Mapping workbook (EFRAG + GRI,
Nov 2024) into a structured hop-2 gold table: for each ESRS datapoint,
which GRI Standard/Disclosure/sub-point it maps to (if any).

This is your GRI->ESRS anchor for the transitive BRSR->GRI->ESRS
composition. It's a much cleaner and finer-grained source than the
Interoperability Index PDF (datapoint-level, not just disclosure-level),
so treat this xlsx as primary and the PDF as a secondary cross-check if
you get to it.

One useful side effect: this workbook has an explicit "Data Type" column
(narrative / percent / monetary / etc.) per ESRS datapoint. That's exactly
the kind of unit/datatype metadata that was missing on the ESRS side in
your earlier ontology build and caused the property_score=0.5 constant bug
we found in mapping_summary.csv. When you populate the ESRS ontology
nodes, pull esrs_data_type from this table -- don't leave it as a
placeholder again.

Sheets skipped: "About this GRI tool", "Index" (front matter, no data).

Column layout varies slightly by sheet (e.g. "ESRS 2 MDR" has one fewer
Appendix column than the topical standard sheets), so columns are located
by header text, not fixed position -- don't hardcode column letters here
even if it looks like it'd save a line.

Output: data/interim/gri_esrs_datapoint_mapping.csv
"""

import csv
import pathlib
import openpyxl

BASE = pathlib.Path(__file__).resolve().parents[1]
XLSX_PATH = BASE / "data" / "raw" / "esrs-gri-standards-data-point-mapping.xlsx"
OUT_PATH = BASE / "data" / "interim" / "gri_esrs_datapoint_mapping.csv"

SKIP_SHEETS = {"About this GRI tool", "Index"}

# canonical column keys we need, mapped from the header text that identifies them
COLUMN_HEADER_MAP = {
    "esrs_datapoint_id": "ID",
    "esrs_topic_code": "ESRS",
    "esrs_dr": "DR",
    "esrs_paragraph": "Paragraph",
    "esrs_name": "Name",          # NOTE: "Name" appears twice (ESRS + GRI side);
    "esrs_data_type": "Data Type",
    "gri_standard": "Standard",
    "gri_disclosure": "Disclosure",
    "gri_number": "Number",
    "gri_name": "Name",           # handled specially below (second occurrence)
    "notes": "Notes",
}


def find_header_row(ws, max_scan_rows=5):
    for r in range(1, max_scan_rows + 1):
        if ws.cell(row=r, column=1).value == "ID":
            return r
    return None


def build_column_index(ws, header_row):
    """
    Returns a dict of canonical_key -> 1-based column index.
    Handles the "Name" appearing twice (ESRS name = first occurrence,
    GRI name = second occurrence, which comes after the "Standard" column).
    """
    headers = {}
    standard_col = None
    for c in range(1, ws.max_column + 1):
        val = ws.cell(row=header_row, column=c).value
        if val is None:
            continue
        val = str(val).replace("\n", " ").strip()
        if val == "Standard":
            standard_col = c
        headers.setdefault(val, []).append(c)

    idx = {}
    idx["esrs_datapoint_id"] = headers.get("ID", [None])[0]
    idx["esrs_topic_code"] = headers.get("ESRS", [None])[0]
    idx["esrs_dr"] = headers.get("DR", [None])[0]
    idx["esrs_paragraph"] = headers.get("Paragraph", [None])[0]
    idx["esrs_data_type"] = headers.get("Data Type", [None])[0]
    idx["gri_standard"] = standard_col
    idx["gri_disclosure"] = headers.get("Disclosure", [None])[0]
    idx["gri_number"] = headers.get("Number", [None])[0]
    idx["notes"] = headers.get("Notes", [None])[0]

    name_cols = headers.get("Name", [])
    if len(name_cols) >= 1:
        idx["esrs_name"] = name_cols[0]
    if len(name_cols) >= 2:
        idx["gri_name"] = name_cols[1]
    else:
        # fall back: if only one "Name" col found, it's the ESRS-side one;
        # GRI name will be missing for this sheet -- fine, we just won't
        # populate gri_name, downstream code should handle blanks.
        idx["gri_name"] = None

    return idx


def main():
    if not XLSX_PATH.exists():
        raise FileNotFoundError(
            f"{XLSX_PATH} not found. Copy esrs-gri-standards-data-point-mapping.xlsx "
            f"into data/raw/ first."
        )

    wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)
    records = []
    skipped_rows = 0

    for sheet_name in wb.sheetnames:
        if sheet_name in SKIP_SHEETS:
            continue
        ws = wb[sheet_name]
        header_row = find_header_row(ws)
        if header_row is None:
            print(f"WARNING: no header row found in sheet '{sheet_name}', skipping")
            continue
        col = build_column_index(ws, header_row)

        for r in range(header_row + 1, ws.max_row + 1):
            dp_id = ws.cell(row=r, column=col["esrs_datapoint_id"]).value
            if not dp_id:
                skipped_rows += 1
                continue

            def get(key):
                c = col.get(key)
                if c is None:
                    return None
                v = ws.cell(row=r, column=c).value
                return str(v).strip() if v is not None else None

            gri_standard = get("gri_standard")
            records.append({
                "esrs_datapoint_id": dp_id,
                "esrs_sheet": sheet_name,
                "esrs_topic_code": get("esrs_topic_code"),
                "esrs_dr": get("esrs_dr"),
                "esrs_paragraph": get("esrs_paragraph"),
                "esrs_name": get("esrs_name"),
                "esrs_data_type": get("esrs_data_type"),
                "gri_standard": gri_standard,
                "gri_disclosure": get("gri_disclosure"),
                "gri_number": get("gri_number"),
                "gri_name": get("gri_name"),
                "notes": get("notes"),
                "has_gri_mapping": bool(
                    gri_standard
                    and (
                        get("gri_disclosure")
                        or get("gri_number")
                    )
                ),
            })

    OUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    with open(OUT_PATH, "w", newline="", encoding="utf-8") as f:
        fieldnames = [
            "esrs_datapoint_id", "esrs_sheet", "esrs_topic_code", "esrs_dr",
            "esrs_paragraph", "esrs_name", "esrs_data_type", "gri_standard",
            "gri_disclosure", "gri_number", "gri_name", "notes", "has_gri_mapping",
        ]
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(records)

    n_mapped = sum(1 for r in records if r["has_gri_mapping"])
    print(f"Parsed {len(records)} ESRS datapoints -> {OUT_PATH}")
    print(f"  {n_mapped} have a GRI mapping, {len(records) - n_mapped} do not")
    print(f"  {skipped_rows} blank/instruction rows skipped")
    print(
        "\nSpot-check a few rows against the source xlsx by eye before trusting "
        "this. In particular check the 'ESRS 2' and 'ESRS 2 MDR' sheets since "
        "they have a slightly different column layout than the topical (E1-E5, "
        "S1-S4, G1) sheets -- that's exactly the kind of thing the dynamic "
        "column-detection is meant to handle, but verify it actually did."
    )


if __name__ == "__main__":
    main()
